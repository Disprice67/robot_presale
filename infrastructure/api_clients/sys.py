import aiohttp
import asyncio
from datetime import datetime
from typing import Optional
from pathlib import Path
from openpyxl import Workbook
import pandas as pd
from playwright.async_api import async_playwright
from settings.config import SysData, HuaweiHeader
from core import IRobotLogger
from io import BytesIO
import base64


class ParsingSYS:
    TAKE = 80
    LOGIN_URL = "https://awsservice.croc.ru/"

    def __init__(self, settings_sys: SysData, header: HuaweiHeader, robot_logger: IRobotLogger):
        self.url = settings_sys.url_sys_agreements
        self.username = settings_sys.username
        self.password = settings_sys.password
        self.robot_logger = robot_logger
        self.headers = {'User-Agent': header.user_agent}
        self.access_token = None
        self.params = {
            "columns": [
                {
                    "field": "redirectCol",
                    "title": "",
                    "width": 40,
                    "sortable": False,
                    "isVisible": True,
                    "displayFieldName": None,
                    "excelExportName": None,
                    "isIconColumn": True
                },
                {
                    "field": "crocCode",
                    "title": "КрокКод СДО",
                    "width": 150,
                    "sortable": True,
                    "isVisible": True,
                    "displayFieldName": None,
                    "excelExportName": "CrocCode",
                    "isIconColumn": False
                },
                {
                    "field": "hierarchicalNumber",
                    "title": "Иерархический номер",
                    "width": 200,
                    "sortable": True,
                    "isVisible": True,
                    "displayFieldName": None,
                    "excelExportName": "HierarchicalNumber",
                    "isIconColumn": False
                },
                {
                    "field": "organization.shortName",
                    "title": "Клиент",
                    "width": 300,
                    "sortable": True,
                    "isVisible": True,
                    "displayFieldName": "organization.displayName",
                    "excelExportName": "OrganizationName",
                    "isIconColumn": False
                },
                {
                    "field": "startDate",
                    "title": "Дата начала",
                    "width": 100,
                    "sortable": True,
                    "isVisible": True,
                    "displayFieldName": None,
                    "excelExportName": "StartDate",
                    "isIconColumn": False
                },
                {
                    "field": "endDate",
                    "title": "Дата окончания",
                    "width": 125,
                    "sortable": True,
                    "isVisible": True,
                    "displayFieldName": None,
                    "excelExportName": "EndDate",
                    "isIconColumn": False
                },
                {
                    "field": "activity.name",
                    "title": "Активность",
                    "width": 300,
                    "sortable": True,
                    "isVisible": True,
                    "displayFieldName": "activity.displayName",
                    "excelExportName": "ActivityName",
                    "isIconColumn": False
                },
                {
                    "field": "status",
                    "title": "Статус",
                    "width": 100,
                    "sortable": True,
                    "isVisible": True,
                    "displayFieldName": None,
                    "excelExportName": "Status",
                    "isIconColumn": False
                },
                {
                    "field": "executor.LastName",
                    "title": "Администратор",
                    "width": 200,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": "administrator",
                    "excelExportName": "AdministratorName",
                    "isIconColumn": False
                },
                {
                    "field": "currencyCode",
                    "title": "Валюта документа",
                    "width": 160,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "CurrencyCode",
                    "isIconColumn": False
                },
                {
                    "field": "prolongationAgreement",
                    "title": "Договор пролонгации",
                    "width": 180,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "ProlongationAgreement",
                    "isIconColumn": False
                },
                {
                    "field": "comment",
                    "title": "Комментарий",
                    "width": 260,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "Comment",
                    "isIconColumn": False
                },
                {
                    "field": "isRequiredNotifications",
                    "title": "Нотификация по заявкам",
                    "width": 200,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "IsRequiredNotifications",
                    "isIconColumn": False
                },
                {
                    "field": "refusalCause",
                    "title": "Причина отказа от пролонгации",
                    "width": 240,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": "refusalCause.displayName",
                    "excelExportName": "RefusalCauseName",
                    "isIconColumn": False
                },
                {
                    "field": "prolongationType",
                    "title": "Пролонгация",
                    "width": 120,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "ProlongationType",
                    "isIconColumn": False
                },
                {
                    "field": "serviceManagers",
                    "title": "Сервис-менеджеры",
                    "width": 300,
                    "sortable": False,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "ServiceManagers",
                    "isIconColumn": False
                },
                {
                    "field": "projectManagers",
                    "title": "Менеджеры проекта",
                    "width": 300,
                    "sortable": False,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "ProjectManagers",
                    "isIconColumn": False
                },
                {
                    "field": "techManagers",
                    "title": "Технические менеджеры",
                    "width": 300,
                    "sortable": False,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "TechManagers",
                    "isIconColumn": False
                },
                {
                    "field": "documentumFileLink",
                    "title": "Ссылка на документ с ЕЦП",
                    "width": 300,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "DocumentumFileLink",
                    "isIconColumn": False
                },
                {
                    "field": "amount",
                    "title": "Сумма контракта в валюте документа",
                    "width": 280,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "Amount",
                    "isIconColumn": False
                },
                {
                    "field": "amountRub",
                    "title": "Сумма контракта в рублях",
                    "width": 200,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "AmountRub",
                    "isIconColumn": False
                },
                {
                    "field": "firm.name",
                    "title": "Фирма",
                    "width": 90,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": "firm.displayName",
                    "excelExportName": "FirmName",
                    "isIconColumn": False
                },
                {
                    "field": "directions",
                    "title": "Направления договора",
                    "width": 300,
                    "sortable": False,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "Directions",
                    "isIconColumn": False
                },
                {
                    "field": "directors",
                    "title": "Директор Клиента",
                    "width": 300,
                    "sortable": False,
                    "isVisible": True,
                    "displayFieldName": None,
                    "excelExportName": "DirectorName",
                    "isIconColumn": False
                },
                {
                    "field": "organization.currentClientSegment",
                    "title": "Сегмент клиента ДИРС2",
                    "width": 200,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "CurrentClientSegment",
                    "isIconColumn": False
                },
                {
                    "field": "organization.clientCardIB.SegmentManual",
                    "title": "Сегмент клиента ИБ1",
                    "width": 200,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": "currentClientSegmentIB",
                    "excelExportName": "CurrentClientSegmentIB",
                    "isIconColumn": False
                },
                {
                    "field": "servicePlans",
                    "title": "Программы обслуживания",
                    "width": 200,
                    "sortable": False,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "ServicePlans",
                    "isIconColumn": False
                },
                {
                    "field": "subAgreements",
                    "title": "Субподрядные договоры",
                    "width": 200,
                    "sortable": False,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "SubAgreements",
                    "isIconColumn": False
                },
                {
                    "field": "otherAgreementForCU.crocCode",
                    "title": "Договор учета КЕ",
                    "width": 200,
                    "sortable": True,
                    "isVisible": False,
                    "displayFieldName": "otherAgreementForCU.displayName",
                    "excelExportName": "OtherAgreementForCU",
                    "isIconColumn": False
                },
                {
                    "field": "isOtherClientAgreements",
                    "title": "По Клиенту есть договоры других направлений",
                    "width": 200,
                    "sortable": False,
                    "isVisible": False,
                    "displayFieldName": None,
                    "excelExportName": "IsOtherClientAgreements",
                    "isIconColumn": False
                }
            ],
            "filter": {
                "directions": {
                    "IsListSearch": True,
                    "ItemList": ["DIR000094"]
                },
                "negativeFilters": [],
                "onlyService": False,
                "gridState": {
                    "skip": 0,
                    "take": 80,
                    "sort": []
                }
            }
        }

    async def _refresh_token(self) -> bool:
        """Эмулирует вход через Playwright, ожидая авторизацию, и извлекает токен или cookies."""
        async with async_playwright() as p:
            try:
                browser = await p.chromium.launch(
                    headless=False,
                    args=['--ignore-certificate-errors', '--incognito']
                )
                
                context = await browser.new_context(
                    ignore_https_errors=True,
                    no_viewport=False,
                    http_credentials={
                        'username': self.username,
                        'password': self.password
                    }
                )
                
                page = await context.new_page()

                self.robot_logger.debug("Открываем страницу логина в режиме инкогнито")
                await page.goto(self.LOGIN_URL, wait_until="domcontentloaded", timeout=60000)

                await asyncio.sleep(3)

                current_url = page.url
                self.robot_logger.debug(f"Текущий URL после аутентификации: {current_url}")

                cookies = await context.cookies()
                self.robot_logger.debug(f"Найденные cookies: {[c['name'] for c in cookies]}")
                
                for cookie in cookies:
                    if cookie["name"] in ["ADFS.OIDC.Token"]:
                        self.access_token = cookie["value"]
                        self.robot_logger.success(f"Токен {cookie['name']} успешно извлечен")
                        await browser.close()
                        return True

                self.robot_logger.error("Не удалось найти токен аутентификации")
                await browser.close()
                return False

            except Exception as e:
                self.robot_logger.critical(f"Ошибка при получении токена: {e}")
                return False

    async def _check_token_status(self) -> bool:
        """Проверяет наличие токена и обновляет его, если необходимо."""
        if self.access_token:
            self.robot_logger.debug("Токен валиден")
            return True

        self.robot_logger.debug("Токен отсутствует, обновляем")
        return await self._refresh_token()

    async def _get_cookies(self) -> dict:
        """Формирует словарь cookies только с основным токеном."""
        if not await self._check_token_status():
            self.robot_logger.critical("Не удалось получить валидный токен")
            raise ValueError("Токен недоступен")
        
        return {"ADFS.OIDC.Token": self.access_token}

    async def _post(self, url: str) -> Optional[dict]:
        """Отправляет асинхронный POST-запрос только с основным токеном в cookies."""
        try:
            headers = {
                "User-Agent": self.headers["User-Agent"],
                "Content-Type": "application/json"
            }
            
            cookies = await self._get_cookies()
            
            async with aiohttp.ClientSession() as session:
                async with session.post(
                    url=url,
                    headers=headers,
                    cookies=cookies,
                    json=self.params,
                    ssl=False
                ) as response:
                    response.raise_for_status()
                    return await response.json()
                    
        except aiohttp.ClientError as e:
            if getattr(e, 'status', None) == 401:
                self.robot_logger.debug("Cookie недействителен, пытаемся обновить")
                self.access_token = None
                if await self._refresh_token():
                    return await self._post(url)
            self.robot_logger.error(f"Ошибка при выполнении запроса: {e}")
            return None

    def _decode_and_process_file(self, encoded_data: str, output_path: Path) -> None:
        """Декодирует файл, извлекает уникальные активные коды и сохраняет их в Excel."""
        try:
            decoded_data = base64.b64decode(encoded_data)
            with pd.ExcelFile(BytesIO(decoded_data)) as xls:
                df = pd.read_excel(xls, sheet_name=0)

            codes = {
                active[1:9] for active in df.get('Активность', [])
                if isinstance(active, str) and len(active) > 8
            }

            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'КОД ПРОЕКТА'
            for ind, code in enumerate(codes, start=2):
                ws[f'A{ind}'] = code

            wb.save(output_path)
            self.robot_logger.success('Договора из СУС обновлены.')
        except Exception as e:
            self.robot_logger.error(f"Ошибка при декодировании и обработке файла: {e}")

    async def parsing_active(self, sys_dir: Path) -> None:
        """Выполняет асинхронный парсинг активных кодов и сохраняет их в файл Excel."""
        self.robot_logger.debug('Процесс обновления договоров через СУС')
        fulldir = sys_dir / 'Договора.xlsx'
        response_data = await self._post(self.url)
        if not response_data:
            return

        self._decode_and_process_file(
            response_data['data']['fileContent'],
            fulldir,
        )