import pandas as pd
from pathlib import Path
from core import DataGenerate, InputData, IExcelHandler, IRobotLogger
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from typing import Optional


class ExcelHandler(IExcelHandler):
    _cell_style_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    _cell_style_fill = PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type="solid"
    )

    _formules = {
        'PRICE/USD': '=IF(U{row}="","",U{row}*2+T{row})',
        'СТОИМОСТЬ ДОСТАВКИ/USD': '=IF(U{row}="","",U{row}/2)',
        'СТ-ТЬ ЗИП С НУЛЯ*1,15': '=IF(U{row}="","",(U{row}*2+T{row})*1.15)',
        '10% ОТ РЫН.ЦЕНЫ': '=IF(V{row}="","",O{row}*E{row}*0.1)',
        'РУБ, СТОИМОСТЬ ПОДДЕРЖКИ': '=IF(E{row}="","",E{row}*F{row})',
        'HOURS': '=IF(E{row}="","",E{row}*G{row})'
    }

    def __init__(self, buffer: Path, robot_logger: IRobotLogger):
        self._buffer = buffer
        self.robot_logger = robot_logger

    @property
    def _sample_file(self,):
        return self._buffer / 'write_sample' / 'sample.xlsx'

    def get_output_file(self, filename: str):
        return self._buffer / 'out' / f'{filename}'

    def read_excel(self, file_path_in: Path) -> Optional[list[InputData]]:
        """get_data_input."""
        mass: list = []
        try:
            with pd.ExcelFile(file_path_in) as excel_file:
                sheetnames = excel_file.sheet_names
                for sheet in sheetnames:
                    if sheet not in ('Оценка рыночной стоимости', 'Для архива'):
                        df_excel = pd.read_excel(excel_file, sheet, na_filter=False)
                        df_excel.columns = df_excel.columns.str.upper()
                        records = df_excel.to_dict('records')
                        a_data = DataGenerate(input_data=records, sheet_name=sheet)
                        list.append(mass, a_data)
            return mass
        except Exception as e:
            self.robot_logger.error(f'Ошибка при обработке/чтение файла Input_Excel {e}')
            return

    def _apply_style_to_range(self, ws: Worksheet, start_cell, end_cell, border=None):
        """Применяет стиль к диапазону ячеек."""
        for row in ws[start_cell:end_cell]:
            for cell in row:
                cell.border = border

    def _record_page_calculation(self, ws_calculation: Worksheet, data: dict, columns: dict, row: int):
        """Записывает данные на страницу 'Расчет'."""
        for key, value in data.items():
            if key in columns:
                col = columns[key]
                ws_calculation.cell(row=row, column=col, value=value)
                if 'MATCH_TYPE' in data:
                    ws_calculation.cell(row=row, column=columns['ЗИП']).fill = self._cell_style_fill

        for key, formula in self._formules.items():
            if key in columns:
                col = columns[key]
                formula_with_row = formula.format(row=row)
                ws_calculation.cell(row=row, column=col, value=formula_with_row)

    def _style_page_calculation(self, ws_calculation: Worksheet):
        """Добавляет стили на страницу 'Расчет'."""
        if ws_calculation.max_row == 0 or ws_calculation.max_column == 0:
            return
        last_column_letter = get_column_letter(ws_calculation.max_column)

        ws_calculation.auto_filter.ref = f"A1:{last_column_letter}1"
        self._apply_style_to_range(
            ws_calculation, 'A1',
            f'{last_column_letter}{ws_calculation.max_row}',
            border=self._cell_style_border
        )

    def _record_page_archive(self, ws_archive: Worksheet, row: int):
        """Записывате данные на страницу 'Для архива'."""
        new_row = []
        for cell in ws_archive[2]:
            new_cell = ws_archive.cell(row=row, column=cell.column)
            new_cell.value = cell.value
            new_row.append(new_cell)
        for cell in new_row:
            if cell.data_type == 'f':
                new_formula = cell.value.replace('2', str(row))
                cell.value = new_formula
            ws_archive.cell(row=row, column=cell.column, value=cell.value).border = self._cell_style_border

    def write_to_excel(self, data: dict, filename: str):
        """Записывает данные в файл Excel с обработкой исключений и логированием."""
        try:
            wb = load_workbook(self._sample_file)
            ws_calculation = wb['Расчет']
            ws_archive = wb['Для архива']
        except FileNotFoundError:
            error_message = f"Файл-шаблон '{self._sample_file}' не найден."
            self.robot_logger.error(error_message)
            return
        except KeyError:
            error_message = "Шаблон не содержит листа 'Расчет'."
            self.robot_logger.error(error_message)
            return

        try:
            row = 2
            columns = {cell.value: cell.column for cell in ws_calculation[1] if cell.value}
            for item in data['input_data']:
                data_dict_upper = {key.upper(): value for key, value in item.items()}
                self._record_page_calculation(ws_calculation, data_dict_upper, columns, row)
                self._record_page_archive(ws_archive, row)
                row += 1
            self._style_page_calculation(ws_calculation)
            output_file = self.get_output_file(filename)
            wb.save(output_file)
            self.robot_logger.success(f"Файл '{filename}' успешно создан.")
            return True
        except Exception as e:
            error_message = f"Ошибка при обработке или записи данных'{filename}': {e}"
            self.robot_logger.error(error_message)
            return
